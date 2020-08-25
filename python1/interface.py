# -*- coding:utf-8 -*-
# @Time  : 2020/8/21  18:34
# @Author  :Jiege
# @Email  :770359382@qq.com
# @file  : interface.py
# @Software : PyCharm
import requests
import openpyxl

def get_data(filename, sheet_name):       # 取值打包函数
    excel = openpyxl.load_workbook(filename)
    sheet = excel[sheet_name]
    max_row = sheet.max_row
    list1 = []        # 打包数值的列表
    for item in range(2,max_row+1):  # 按行存入字典
        dict1 = dict(
        id = sheet.cell(row=item, column=1).value,
        url = sheet.cell(row=item, column=5).value,
        data = sheet.cell(row=item, column=6).value,
        expected = sheet.cell(row=item, column=7).value)
        list1.append(dict1)
    return list1

filename = 'test_case_api.xlsx'
sheet_name = 'login'             # 把文件名和sheet名提出来，不局限与get函数
package = get_data(filename, sheet_name)  # 得到的便是打包数据，一个列表
headers = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}   # 请求头

def request(url, data):  #请求函数
    url = url
    json = data
    reg = requests.post(url=url, json=json, headers=headers).json()
    return reg

def case(filename, sheet_name):
    for text_case in package:  # 数据列表拆出来
        excel = openpyxl.load_workbook(filename)
        sheet = excel[sheet_name]
        id = text_case['id']
        url = text_case['url']
        data = eval(text_case['data'])   # 如果是充值，加一个memberid
        expected = eval(text_case['expected'])  # 预期结果
        msg = request(url, data)               # 实际结果，请求函数执行后的响应结果
        print('用例id：{}'.format(id) + '\n' + '预期结果为：{}'.format(expected['msg']) + '\n' + '实际结果为：{}'.format(msg['msg']))
        if expected['msg'] == msg['msg']:  # 比较预期和实际结果的msg
            print('预期与实际一致，此用例通过。')
            sheet.cell(row=id+1, column=8).value = '是'
        else:
            print('预期与实际不同，此用例不通过！')
            sheet.cell(row=id+1, column=8).value = '否'
        excel.save(filename)
        print('*' * 100)
case(filename, sheet_name)
